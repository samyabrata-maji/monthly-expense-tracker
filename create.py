import os
import csv
import openpyxl
from datetime import date


# uses date as file name
# eg Nov_02_2022.xlsx for entries on Nov 2nd
filename = date.today().strftime('%b_%d_%y')
parent_dir = fr"{os.getcwd()}\data"
filepath_excel = fr'{parent_dir}\{filename}.xlsx'
filepath_csv = fr'{parent_dir}\{filename}.csv'


def import_from_csv(worksheet, workbook):
    # if no csv files exist then creates one
    if not os.path.exists(filepath_csv):
        open(filepath_csv, 'x')

    # overwrites data from csv to excel
    with open(filepath_csv, 'w') as csv_file:
        reader = csv.reader(csv_file, delimiter=':')
        for row in reader:
            worksheet.append(row)
    workbook.save('file.xlsx')


# if no file with the specific filename exists, then creates one
# otherwise loads the file
if not os.path.exists(filepath_excel):
    wb = openpyxl.Workbook()
    ws = wb.create_sheet()
    wb.save(filename=filepath_excel)
    print(f'File {filename}.xlsx created successfully')
else:
    print('File already created')
    wb = openpyxl.load_workbook(filename=filepath_excel)
    ws = wb.active

# import data from CSV file to the Excel file of the same name
if input('Import CSV data? (y/n)') == 'y':
    print('importing...')
    import_from_csv(workbook=wb, worksheet=ws)

wb.close()
